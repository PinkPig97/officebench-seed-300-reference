#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
REFERENCE_ANSWER = PACKAGE_ROOT / "answer" / "gold_submission.xlsx"
CELL_RE = re.compile(r"^\$?([A-Z]+)\$?(\d+)$")
NUMBER_RE = re.compile(r"^\d+(?:\.\d+)?$")
MANIFEST = json.loads(
    """
{
  "requiredSheets": [
    "Inputs",
    "Debt_Schedule",
    "Covenant_Test",
    "Summary"
  ],
  "headerCells": {
    "Inputs!B9": "2026Q1",
    "Debt_Schedule!B2": "2026Q1",
    "Covenant_Test!B2": "2026Q1",
    "Summary!B10": "2026Q1",
    "Inputs!C9": "2026Q2",
    "Debt_Schedule!C2": "2026Q2",
    "Covenant_Test!C2": "2026Q2",
    "Summary!C10": "2026Q2",
    "Inputs!D9": "2026Q3",
    "Debt_Schedule!D2": "2026Q3",
    "Covenant_Test!D2": "2026Q3",
    "Summary!D10": "2026Q3",
    "Inputs!E9": "2026Q4",
    "Debt_Schedule!E2": "2026Q4",
    "Covenant_Test!E2": "2026Q4",
    "Summary!E10": "2026Q4",
    "Inputs!B15": "2025Q1",
    "Inputs!C15": "2025Q2",
    "Inputs!D15": "2025Q3",
    "Inputs!E15": "2025Q4",
    "Inputs!F15": "2026Q1",
    "Inputs!G15": "2026Q2",
    "Inputs!H15": "2026Q3",
    "Inputs!I15": "2026Q4"
  },
  "valueCells": {
    "Inputs!C5": 600,
    "Inputs!C6": 0.08,
    "Inputs!B10": 20,
    "Inputs!B11": 120,
    "Inputs!B12": 30,
    "Debt_Schedule!B4": 600,
    "Debt_Schedule!B5": 20,
    "Debt_Schedule!B6": 12,
    "Debt_Schedule!B7": 580,
    "Covenant_Test!B4": 120,
    "Covenant_Test!B5": 580,
    "Covenant_Test!B6": 460,
    "Covenant_Test!B7": 120,
    "Covenant_Test!B8": 3.8333333333,
    "Summary!B12": 3.8333333333,
    "Inputs!C10": 25,
    "Inputs!C11": 95,
    "Inputs!C12": 34,
    "Debt_Schedule!C4": 580,
    "Debt_Schedule!C5": 25,
    "Debt_Schedule!C6": 11.6,
    "Debt_Schedule!C7": 555,
    "Covenant_Test!C4": 95,
    "Covenant_Test!C5": 555,
    "Covenant_Test!C6": 460,
    "Covenant_Test!C7": 126,
    "Covenant_Test!C8": 3.6507936508,
    "Summary!C12": 3.6507936508,
    "Inputs!D10": 35,
    "Inputs!D11": 80,
    "Inputs!D12": 38,
    "Debt_Schedule!D4": 555,
    "Debt_Schedule!D5": 35,
    "Debt_Schedule!D6": 11.1,
    "Debt_Schedule!D7": 520,
    "Covenant_Test!D4": 80,
    "Covenant_Test!D5": 520,
    "Covenant_Test!D6": 440,
    "Covenant_Test!D7": 134,
    "Covenant_Test!D8": 3.2835820896,
    "Summary!D12": 3.2835820896,
    "Inputs!E10": 40,
    "Inputs!E11": 100,
    "Inputs!E12": 42,
    "Debt_Schedule!E4": 520,
    "Debt_Schedule!E5": 40,
    "Debt_Schedule!E6": 10.4,
    "Debt_Schedule!E7": 480,
    "Covenant_Test!E4": 100,
    "Covenant_Test!E5": 480,
    "Covenant_Test!E6": 380,
    "Covenant_Test!E7": 144,
    "Covenant_Test!E8": 2.6388888889,
    "Summary!E12": 2.6388888889,
    "Inputs!B16": 26,
    "Inputs!C16": 28,
    "Inputs!D16": 30,
    "Inputs!E16": 32,
    "Inputs!F16": 30,
    "Inputs!G16": 34,
    "Inputs!H16": 38,
    "Inputs!I16": 42,
    "Summary!B5": 480,
    "Summary!B6": 2.6388888889
  },
  "labelCells": {
    "Covenant_Test!B9": "BREACH",
    "Summary!B11": "BREACH",
    "Covenant_Test!C9": "BREACH",
    "Summary!C11": "BREACH",
    "Covenant_Test!D9": "WARN",
    "Summary!D11": "WARN",
    "Covenant_Test!E9": "PASS",
    "Summary!E11": "PASS",
    "Summary!B7": "PASS"
  },
  "formulaCells": {
    "Debt_Schedule!B4": "=Inputs!$C$5",
    "Debt_Schedule!C4": "=B7",
    "Debt_Schedule!D4": "=C7",
    "Debt_Schedule!E4": "=D7",
    "Debt_Schedule!B5": "=Inputs!B10",
    "Debt_Schedule!B6": "=B4*Inputs!$C$6/4",
    "Debt_Schedule!B7": "=B4-B5",
    "Covenant_Test!B4": "=Inputs!B11",
    "Covenant_Test!B5": "=Debt_Schedule!B7",
    "Covenant_Test!B6": "=MAX(B5-B4,0)",
    "Summary!B11": "=Covenant_Test!B9",
    "Summary!B12": "=Covenant_Test!B8",
    "Debt_Schedule!C5": "=Inputs!C10",
    "Debt_Schedule!C6": "=C4*Inputs!$C$6/4",
    "Debt_Schedule!C7": "=C4-C5",
    "Covenant_Test!C4": "=Inputs!C11",
    "Covenant_Test!C5": "=Debt_Schedule!C7",
    "Covenant_Test!C6": "=MAX(C5-C4,0)",
    "Summary!C11": "=Covenant_Test!C9",
    "Summary!C12": "=Covenant_Test!C8",
    "Debt_Schedule!D5": "=Inputs!D10",
    "Debt_Schedule!D6": "=D4*Inputs!$C$6/4",
    "Debt_Schedule!D7": "=D4-D5",
    "Covenant_Test!D4": "=Inputs!D11",
    "Covenant_Test!D5": "=Debt_Schedule!D7",
    "Covenant_Test!D6": "=MAX(D5-D4,0)",
    "Summary!D11": "=Covenant_Test!D9",
    "Summary!D12": "=Covenant_Test!D8",
    "Debt_Schedule!E5": "=Inputs!E10",
    "Debt_Schedule!E6": "=E4*Inputs!$C$6/4",
    "Debt_Schedule!E7": "=E4-E5",
    "Covenant_Test!E4": "=Inputs!E11",
    "Covenant_Test!E5": "=Debt_Schedule!E7",
    "Covenant_Test!E6": "=MAX(E5-E4,0)",
    "Summary!E11": "=Covenant_Test!E9",
    "Summary!E12": "=Covenant_Test!E8",
    "Summary!B5": "=Debt_Schedule!E7",
    "Summary!B6": "=Covenant_Test!E8",
    "Summary!B7": "=Covenant_Test!E9"
  },
  "inputValueCells": [
    "Inputs!C5",
    "Inputs!C6",
    "Inputs!B10",
    "Inputs!B11",
    "Inputs!B12",
    "Inputs!C10",
    "Inputs!C11",
    "Inputs!C12",
    "Inputs!D10",
    "Inputs!D11",
    "Inputs!D12",
    "Inputs!E10",
    "Inputs!E11",
    "Inputs!E12",
    "Inputs!B16",
    "Inputs!C16",
    "Inputs!D16",
    "Inputs!E16",
    "Inputs!F16",
    "Inputs!G16",
    "Inputs!H16",
    "Inputs!I16"
  ],
  "modeledFormulaCells": [
    "Debt_Schedule!B4",
    "Debt_Schedule!B5",
    "Debt_Schedule!B6",
    "Debt_Schedule!B7",
    "Covenant_Test!B4",
    "Covenant_Test!B5",
    "Covenant_Test!B6",
    "Covenant_Test!B7",
    "Covenant_Test!B8",
    "Covenant_Test!B9",
    "Summary!B11",
    "Summary!B12",
    "Debt_Schedule!C4",
    "Debt_Schedule!C5",
    "Debt_Schedule!C6",
    "Debt_Schedule!C7",
    "Covenant_Test!C4",
    "Covenant_Test!C5",
    "Covenant_Test!C6",
    "Covenant_Test!C7",
    "Covenant_Test!C8",
    "Covenant_Test!C9",
    "Summary!C11",
    "Summary!C12",
    "Debt_Schedule!D4",
    "Debt_Schedule!D5",
    "Debt_Schedule!D6",
    "Debt_Schedule!D7",
    "Covenant_Test!D4",
    "Covenant_Test!D5",
    "Covenant_Test!D6",
    "Covenant_Test!D7",
    "Covenant_Test!D8",
    "Covenant_Test!D9",
    "Summary!D11",
    "Summary!D12",
    "Debt_Schedule!E4",
    "Debt_Schedule!E5",
    "Debt_Schedule!E6",
    "Debt_Schedule!E7",
    "Covenant_Test!E4",
    "Covenant_Test!E5",
    "Covenant_Test!E6",
    "Covenant_Test!E7",
    "Covenant_Test!E8",
    "Covenant_Test!E9",
    "Summary!E11",
    "Summary!E12",
    "Summary!B5",
    "Summary!B6",
    "Summary!B7"
  ]
}
"""
)


@dataclass(frozen=True)
class CellRef:
    sheet: str
    cell: str


@dataclass(frozen=True)
class RangeRef:
    sheet: str
    start_cell: str
    end_cell: str


class EvaluationError(Exception):
    pass


def normalize_formula(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace(" ", "").replace("$", "").upper()


def split_ref(ref: str) -> tuple[str, str]:
    sheet_name, cell_ref = ref.split("!", 1)
    return sheet_name, cell_ref


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def compare_numeric(actual: Any, expected: Any, tolerance: float) -> bool:
    if actual is None:
        return False
    try:
        return math.isclose(float(actual), float(expected), abs_tol=tolerance, rel_tol=0.0)
    except (TypeError, ValueError):
        return False


def normalize_cell_ref(cell_ref: str) -> str:
    match = CELL_RE.fullmatch(cell_ref.upper())
    if match is None:
        raise EvaluationError(f"unsupported cell reference: {cell_ref}")
    column, row = match.groups()
    return f"{column}{row}"


def col_row_from_cell(cell_ref: str) -> tuple[int, int]:
    normalized = normalize_cell_ref(cell_ref)
    match = CELL_RE.fullmatch(normalized)
    assert match is not None
    column, row = match.groups()
    return column_index_from_string(column), int(row)


def cell_from_col_row(column_index: int, row_index: int) -> str:
    return f"{get_column_letter(column_index)}{row_index}"


def tokenize_formula(expr: str) -> list[str]:
    tokens: list[str] = []
    i = 0
    while i < len(expr):
        ch = expr[i]
        if ch.isspace():
            i += 1
            continue
        if expr.startswith("<=", i) or expr.startswith(">=", i) or expr.startswith("<>", i):
            tokens.append(expr[i : i + 2])
            i += 2
            continue
        if ch in "(),:+-*/!<>=":
            tokens.append(ch)
            i += 1
            continue
        if ch == '"':
            j = i + 1
            while j < len(expr) and expr[j] != '"':
                j += 1
            if j >= len(expr):
                raise EvaluationError(f"unterminated string literal in formula: {expr}")
            tokens.append(expr[i : j + 1])
            i = j + 1
            continue
        j = i
        while j < len(expr) and (expr[j].isalnum() or expr[j] in "._$"):
            j += 1
        if j == i:
            raise EvaluationError(f"unsupported token in formula: {expr[i:]}")
        tokens.append(expr[i:j])
        i = j
    return tokens


class FormulaParser:
    def __init__(self, expr: str, current_sheet: str, evaluator: "WorkbookEvaluator"):
        self.tokens = tokenize_formula(expr)
        self.position = 0
        self.current_sheet = current_sheet
        self.evaluator = evaluator

    def peek(self) -> str | None:
        if self.position >= len(self.tokens):
            return None
        return self.tokens[self.position]

    def consume(self, expected: str | None = None) -> str:
        token = self.peek()
        if token is None:
            raise EvaluationError("unexpected end of formula")
        if expected is not None and token != expected:
            raise EvaluationError(f"expected token {expected}, got {token}")
        self.position += 1
        return token

    def parse(self) -> Any:
        value = self.parse_comparison()
        if self.peek() is not None:
            raise EvaluationError(f"unexpected trailing token: {self.peek()}")
        return value

    def parse_comparison(self) -> Any:
        left = self.parse_additive()
        token = self.peek()
        if token in {"<=", ">=", "<>", "<", ">", "="}:
            operator = self.consume()
            right = self.parse_additive()
            return self.compare_values(left, operator, right)
        return left

    def parse_additive(self) -> Any:
        value = self.parse_multiplicative()
        while self.peek() in {"+", "-"}:
            operator = self.consume()
            right = self.parse_multiplicative()
            left_value = self.evaluator.resolve_scalar(value)
            right_value = self.evaluator.resolve_scalar(right)
            value = left_value + right_value if operator == "+" else left_value - right_value
        return value

    def parse_multiplicative(self) -> Any:
        value = self.parse_range()
        while self.peek() in {"*", "/"}:
            operator = self.consume()
            right = self.parse_range()
            left_value = self.evaluator.resolve_scalar(value)
            right_value = self.evaluator.resolve_scalar(right)
            value = left_value * right_value if operator == "*" else left_value / right_value
        return value

    def parse_range(self) -> Any:
        value = self.parse_unary()
        while self.peek() == ":":
            self.consume(":")
            right = self.parse_unary()
            if not isinstance(value, CellRef) or not isinstance(right, CellRef):
                raise EvaluationError("range endpoints must be cell references")
            if value.sheet != right.sheet:
                if right.sheet == self.current_sheet:
                    right = CellRef(value.sheet, right.cell)
                else:
                    raise EvaluationError("cross-sheet ranges are not supported")
            value = RangeRef(value.sheet, value.cell, right.cell)
        return value

    def parse_unary(self) -> Any:
        token = self.peek()
        if token == "+":
            self.consume("+")
            return self.evaluator.resolve_scalar(self.parse_unary())
        if token == "-":
            self.consume("-")
            return -self.evaluator.resolve_scalar(self.parse_unary())
        return self.parse_primary()

    def parse_primary(self) -> Any:
        token = self.peek()
        if token is None:
            raise EvaluationError("missing primary expression")
        if token == "(":
            self.consume("(")
            value = self.parse_comparison()
            self.consume(")")
            return value
        if token.startswith('"') and token.endswith('"'):
            return self.consume()[1:-1]
        if NUMBER_RE.fullmatch(token):
            raw = self.consume()
            return int(raw) if "." not in raw else float(raw)

        raw = self.consume()
        if self.peek() == "!":
            self.consume("!")
            cell_token = self.consume()
            return CellRef(raw, normalize_cell_ref(cell_token))
        if self.peek() == "(":
            return self.parse_function_call(raw)
        upper = raw.upper()
        if upper == "TRUE":
            return True
        if upper == "FALSE":
            return False
        if CELL_RE.fullmatch(raw.upper()):
            return CellRef(self.current_sheet, normalize_cell_ref(raw))
        raise EvaluationError(f"unsupported token in formula: {raw}")

    def parse_function_call(self, name: str) -> Any:
        self.consume("(")
        args: list[Any] = []
        if self.peek() != ")":
            while True:
                args.append(self.parse_comparison())
                if self.peek() != ",":
                    break
                self.consume(",")
        self.consume(")")
        return self.evaluator.call_function(name, args)

    def compare_values(self, left: Any, operator: str, right: Any) -> bool:
        left_value = self.evaluator.resolve_scalar(left)
        right_value = self.evaluator.resolve_scalar(right)
        if operator == "<=":
            return left_value <= right_value
        if operator == ">=":
            return left_value >= right_value
        if operator == "<":
            return left_value < right_value
        if operator == ">":
            return left_value > right_value
        if operator == "=":
            return left_value == right_value
        if operator == "<>":
            return left_value != right_value
        raise EvaluationError(f"unsupported comparison operator: {operator}")


class WorkbookEvaluator:
    def __init__(self, workbook):
        self.workbook = workbook
        self.memo: dict[tuple[str, str], Any] = {}
        self.in_progress: set[tuple[str, str]] = set()

    def get_cell_value(self, sheet_name: str, cell_ref: str) -> Any:
        normalized_ref = normalize_cell_ref(cell_ref)
        key = (sheet_name, normalized_ref)
        if key in self.memo:
            return self.memo[key]
        if key in self.in_progress:
            raise EvaluationError(f"circular reference detected at {sheet_name}!{normalized_ref}")

        self.in_progress.add(key)
        try:
            raw = self.workbook[sheet_name][normalized_ref].value
            if is_formula(raw):
                parsed = FormulaParser(raw[1:], sheet_name, self).parse()
                value = self.resolve_scalar(parsed)
            else:
                value = raw
            self.memo[key] = value
            return value
        finally:
            self.in_progress.discard(key)

    def resolve_scalar(self, value: Any) -> Any:
        if isinstance(value, CellRef):
            return self.get_cell_value(value.sheet, value.cell)
        if isinstance(value, RangeRef):
            raise EvaluationError("range cannot be used as scalar")
        return value

    def call_function(self, name: str, args: list[Any]) -> Any:
        upper = name.upper()
        if upper == "SUM":
            total = 0.0
            for arg in args:
                total += self.sum_argument(arg)
            return total
        if upper == "MAX":
            values = [self.resolve_scalar(arg) for arg in args]
            return max(values)
        if upper == "IF":
            if len(args) != 3:
                raise EvaluationError("IF expects 3 arguments")
            return args[1] if self.resolve_scalar(args[0]) else args[2]
        if upper == "MATCH":
            if len(args) != 3:
                raise EvaluationError("MATCH expects 3 arguments")
            lookup_value = self.resolve_scalar(args[0])
            match_range = self.ensure_range(args[1])
            match_type = self.resolve_scalar(args[2])
            if int(match_type) != 0:
                raise EvaluationError("only exact MATCH(..., ..., 0) is supported")
            for index, candidate in enumerate(self.range_values(match_range), start=1):
                if candidate == lookup_value:
                    return index
            raise EvaluationError(f"MATCH could not find {lookup_value}")
        if upper == "INDEX":
            if len(args) != 3:
                raise EvaluationError("INDEX expects 3 arguments")
            source_range = self.ensure_range(args[0])
            row_num = int(self.resolve_scalar(args[1]))
            col_num = int(self.resolve_scalar(args[2]))
            return self.range_index(source_range, row_num, col_num)
        raise EvaluationError(f"unsupported function: {name}")

    def ensure_range(self, value: Any) -> RangeRef:
        if isinstance(value, RangeRef):
            return value
        if isinstance(value, CellRef):
            return RangeRef(value.sheet, value.cell, value.cell)
        raise EvaluationError("expected range reference")

    def sum_argument(self, value: Any) -> float:
        if isinstance(value, RangeRef):
            return sum(float(item) for item in self.range_values(value))
        return float(self.resolve_scalar(value))

    def range_values(self, range_ref: RangeRef) -> list[Any]:
        start_col, start_row = col_row_from_cell(range_ref.start_cell)
        end_col, end_row = col_row_from_cell(range_ref.end_cell)
        values: list[Any] = []
        for row_index in range(min(start_row, end_row), max(start_row, end_row) + 1):
            for col_index in range(min(start_col, end_col), max(start_col, end_col) + 1):
                values.append(self.get_cell_value(range_ref.sheet, cell_from_col_row(col_index, row_index)))
        return values

    def range_index(self, range_ref: RangeRef, row_num: int, col_num: int) -> CellRef:
        start_col, start_row = col_row_from_cell(range_ref.start_cell)
        end_col, end_row = col_row_from_cell(range_ref.end_cell)
        row_count = max(start_row, end_row) - min(start_row, end_row) + 1
        col_count = max(start_col, end_col) - min(start_col, end_col) + 1
        if not (1 <= row_num <= row_count and 1 <= col_num <= col_count):
            raise EvaluationError("INDEX out of range")
        resolved_row = min(start_row, end_row) + row_num - 1
        resolved_col = min(start_col, end_col) + col_num - 1
        return CellRef(range_ref.sheet, cell_from_col_row(resolved_col, resolved_row))


def numeric_tolerance(ref: str) -> float:
    if ref == "Inputs!C6":
        return 0.0001
    if ref.endswith(("!B8", "!C8", "!D8", "!E8")) or ref.endswith(("!B12", "!C12", "!D12", "!E12")) or ref == "Summary!B6":
        return 0.01
    return 0.01


def build_failure_payload(submission_path: Path, failures: list[str], total_checks: int, passed_checks: int, check_results: dict[str, dict[str, int]]) -> dict[str, Any]:
    score = 1.0 if total_checks == 0 else round(passed_checks / total_checks, 4)
    return {
        "task_id": "seed_300",
        "task_pass": len(failures) == 0,
        "score": score,
        "total_checks": total_checks,
        "passed_checks": passed_checks,
        "failures_count": len(failures),
        "failures": failures,
        "check_results": check_results,
        "submission": str(submission_path),
        "reference_answer": str(REFERENCE_ANSWER),
    }


def run_checks(submission_path: Path) -> dict[str, Any]:
    wb_values = load_workbook(submission_path, data_only=True)
    wb_formulas = load_workbook(submission_path, data_only=False)
    evaluator = WorkbookEvaluator(wb_formulas)

    failures: list[str] = []
    total_checks = 0
    passed_checks = 0
    check_results: dict[str, dict[str, int]] = {}

    def record(category: str, ok: bool, message: str | None = None) -> None:
        nonlocal total_checks, passed_checks
        bucket = check_results.setdefault(category, {"passed": 0, "total": 0})
        bucket["total"] += 1
        total_checks += 1
        if ok:
            bucket["passed"] += 1
            passed_checks += 1
        elif message:
            failures.append(message)

    def get_actual_value(ref: str) -> tuple[Any, str | None]:
        sheet_name, cell_ref = split_ref(ref)
        cached_value = wb_values[sheet_name][cell_ref].value
        if cached_value is not None:
            return cached_value, None
        try:
            return evaluator.get_cell_value(sheet_name, cell_ref), None
        except EvaluationError as exc:
            return None, f"evaluation error at {ref}: {exc}"

    for sheet_name in MANIFEST["requiredSheets"]:
        record("required_sheets", sheet_name in wb_values.sheetnames, f"missing required sheet: {sheet_name}")

    missing_sheet = any(sheet_name not in wb_values.sheetnames for sheet_name in MANIFEST["requiredSheets"])
    if missing_sheet:
        return build_failure_payload(submission_path, failures, total_checks, passed_checks, check_results)

    for ref, expected in MANIFEST["headerCells"].items():
        actual, error = get_actual_value(ref)
        if error:
            record("formula_evaluation", False, error)
            continue
        record("header_cells", actual == expected, f"header mismatch at {ref}: expected {expected}, got {actual}")

    for ref, expected in MANIFEST["valueCells"].items():
        actual, error = get_actual_value(ref)
        if error:
            record("formula_evaluation", False, error)
            continue
        record(
            "value_cells",
            compare_numeric(actual, expected, numeric_tolerance(ref)),
            f"value mismatch at {ref}: expected {expected}, got {actual}",
        )

    for ref, expected in MANIFEST["labelCells"].items():
        actual, error = get_actual_value(ref)
        if error:
            record("formula_evaluation", False, error)
            continue
        record("label_cells", actual == expected, f"label mismatch at {ref}: expected {expected}, got {actual}")

    for ref, expected in MANIFEST["formulaCells"].items():
        sheet_name, cell_ref = split_ref(ref)
        actual = wb_formulas[sheet_name][cell_ref].value
        record(
            "formula_cells",
            normalize_formula(actual) == normalize_formula(expected),
            f"formula mismatch at {ref}: expected {expected}, got {actual}",
        )

    for ref in MANIFEST["inputValueCells"]:
        sheet_name, cell_ref = split_ref(ref)
        actual = wb_formulas[sheet_name][cell_ref].value
        record(
            "input_value_cells",
            not is_formula(actual),
            f"input cell must be value, not formula: {ref} got {actual}",
        )

    for ref in MANIFEST["modeledFormulaCells"]:
        sheet_name, cell_ref = split_ref(ref)
        actual = wb_formulas[sheet_name][cell_ref].value
        record(
            "modeled_formula_cells",
            is_formula(actual),
            f"modeled cell must be formula: {ref} got {actual}",
        )

    return build_failure_payload(submission_path, failures, total_checks, passed_checks, check_results)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--submission", required=True, type=Path, help="Path to the submission workbook")
    args = parser.parse_args()

    submission_path = args.submission.expanduser().resolve()
    if not submission_path.is_file():
        payload = build_failure_payload(
            submission_path,
            [f"submission not found: {submission_path}"],
            total_checks=0,
            passed_checks=0,
            check_results={},
        )
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return 1

    result = run_checks(submission_path)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["task_pass"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
